"""wt_pricesync_report.py — گزارشِ اکسلِ مرتبِ بروزرسانی‌های قیمت/موجودی (سینکِ کانالِ CAT Group → سایت).

خروجی: یک فایلِ .xlsx راست‌به‌چپ با شیت‌های خلاصه/تغییراتِ قیمت/تغییراتِ موجودی/هشدارها — مناسبِ ارسال به پیویِ اپراتور.
plan = خروجیِ wt_pricesync.plan_changes ؛ products = list از product_view (برای نام محصول)؛
apply_result = خروجیِ apply_plan (اختیاری) تا خطاها علامت بخورند.
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:                       # تاریخِ جلالی اگر در دسترس بود
    import clock
except Exception:          # noqa: BLE001
    clock = None

_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
_UP_FILL = PatternFill("solid", fgColor="FCE4D6")     # قیمتِ بالارونده (نارنجیِ کم‌رنگ)
_DOWN_FILL = PatternFill("solid", fgColor="E2EFDA")   # قیمتِ پایین‌رونده (سبزِ کم‌رنگ)
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_ERR_FONT = Font(color="C00000", bold=True)
_MONEY = "#,##0"
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def _now_jalali() -> str:
    if clock is not None:
        for fn in ("jalali_str", "jalali_now", "now_jalali"):
            f = getattr(clock, fn, None)
            if callable(f):
                try:
                    return str(f())
                except Exception:  # noqa: BLE001
                    pass
    return ""


def _header(ws, titles, row=1):
    for i, t in enumerate(titles, 1):
        c = ws.cell(row=row, column=i, value=t)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(plan: dict, products: list, meta: dict, out_path: str,
                applied: bool = False, apply_result: dict | None = None) -> str:
    by_id = {p.get("id"): p for p in (products or [])}

    def nm(pid):
        v = by_id.get(pid)
        return (v.get("name") if v else "") or ""

    err_ids = {e.get("id") for e in ((apply_result or {}).get("errors") or [])}
    mode = "اعمال‌شده" if applied else "پیش‌نمایش (بدونِ نوشتن)"

    wb = Workbook()

    # ---------- شیتِ خلاصه ----------
    ws = wb.active
    ws.title = "خلاصه"
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value=f"گزارشِ سینکِ قیمت/موجودی — {mode}").font = _TITLE_FONT
    dt = _now_jalali()
    ws.cell(row=2, column=1, value=(f"تاریخ: {dt}" if dt else "تاریخ: —"))
    ws.cell(row=3, column=1, value=f"کانال: {meta.get('channel') or '—'} · رفرنسِ خوانده‌شده: {meta.get('count') or 0}")
    rows = [
        ("تغییرِ قیمت (رفرنسِ دقیق)", len(plan["price_exact"])),
        ("تغییرِ قیمت (هم‌خانواده)", len(plan["price_family"])),
        ("→ موجود", len(plan["set_instock"])),
        ("→ ناموجود", len(plan["set_outofstock"])),
        ("موجودیِ تعدادیِ دست‌نخورده", plan["untouched_qty"]),
        ("تعدادیِ بدونِ قیمتِ کانال (بررسی)", len(plan.get("qty_review", []))),
        ("هشدار: مبهمِ هم‌خانواده", len(plan["ambiguous_family"])),
        ("هشدار: رفرنسِ کانال بدونِ محصول", len(plan["unmatched_refs"])),
        ("محصولِ برند بدونِ رفرنس", plan["no_ref"]),
    ]
    if applied:
        rows.append(("خطا در نوشتن", len(err_ids)))
    _header(ws, ["مورد", "تعداد"], row=5)
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=k).border = _BORDER
        c = ws.cell(row=i, column=2, value=v)
        c.alignment = _CENTER
        c.border = _BORDER
    _widths(ws, [38, 12])

    # ---------- شیتِ تغییراتِ قیمت ----------
    wp = wb.create_sheet("تغییرات قیمت")
    wp.sheet_view.rightToLeft = True
    _header(wp, ["ردیف", "رفرنس", "نام محصول", "شناسه", "قیمت قبلی", "قیمت جدید", "اختلاف", "نوع"])
    r = 2
    for kind, items in (("دقیق", plan["price_exact"]), ("هم‌خانواده", plan["price_family"])):
        for c in items:
            diff = c["new"] - c["old"]
            vals = [r - 1, c["ref"], nm(c["id"]), c["id"], c["old"], c["new"], diff, kind]
            for j, v in enumerate(vals, 1):
                cell = wp.cell(row=r, column=j, value=v)
                cell.border = _BORDER
                if j in (5, 6, 7):
                    cell.number_format = _MONEY
                    cell.alignment = _CENTER
            fill = _DOWN_FILL if diff < 0 else _UP_FILL
            for j in (5, 6, 7):
                wp.cell(row=r, column=j).fill = fill
            if c["id"] in err_ids:
                wp.cell(row=r, column=2).font = _ERR_FONT
            r += 1
    _widths(wp, [7, 18, 40, 10, 16, 16, 14, 12])

    # ---------- شیتِ تغییراتِ موجودی ----------
    wsk = wb.create_sheet("تغییرات موجودی")
    wsk.sheet_view.rightToLeft = True
    _header(wsk, ["ردیف", "رفرنس", "نام محصول", "شناسه", "تغییر"])
    r = 2
    for label, items in (("→ موجود", plan["set_instock"]), ("→ ناموجود", plan["set_outofstock"])):
        for c in items:
            vals = [r - 1, c["ref"], nm(c["id"]), c["id"], label]
            for j, v in enumerate(vals, 1):
                cell = wsk.cell(row=r, column=j, value=v)
                cell.border = _BORDER
                if j in (1, 4, 5):
                    cell.alignment = _CENTER
            if c["id"] in err_ids:
                wsk.cell(row=r, column=2).font = _ERR_FONT
            r += 1
    _widths(wsk, [7, 18, 40, 10, 14])

    # ---------- شیتِ هشدارها ----------
    ww = wb.create_sheet("هشدارها")
    ww.sheet_view.rightToLeft = True
    _header(ww, ["نوع", "رفرنس", "شناسه", "توضیح"])
    r = 2
    for c in plan["ambiguous_family"]:
        vals = ["مبهمِ هم‌خانواده", c["ref"], c["id"],
                f"خانواده {c.get('family')} چند قیمت دارد: " + "، ".join(f"{p:,}" for p in c.get("prices", []))]
        for j, v in enumerate(vals, 1):
            cell = ww.cell(row=r, column=j, value=v)
            cell.border = _BORDER
            cell.fill = _WARN_FILL
        r += 1
    for ref in plan["unmatched_refs"]:
        vals = ["رفرنسِ کانال بدونِ محصولِ سایت", ref, "—", "در کانال هست ولی محصولی با این رفرنس روی سایت نیست"]
        for j, v in enumerate(vals, 1):
            cell = ww.cell(row=r, column=j, value=v)
            cell.border = _BORDER
        r += 1
    _widths(ww, [30, 20, 10, 55])

    # ---------- شیتِ محصولاتِ تعدادیِ بدونِ قیمتِ کانال (نیازِ بررسیِ دستی) ----------
    qr = plan.get("qty_review") or []
    wq = wb.create_sheet("تعدادیِ بی‌قیمت")
    wq.sheet_view.rightToLeft = True
    _header(wq, ["ردیف", "رفرنس", "نام محصول", "شناسه", "قیمتِ فعلی", "خانواده", "وضعیت"])
    for i, c in enumerate(qr, 1):
        vals = [i, c.get("ref"), c.get("name") or "", c.get("id"), int(c.get("price", 0) or 0),
                c.get("family") or "—", c.get("reason") or ""]
        for j, v in enumerate(vals, 1):
            cell = wq.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
            cell.fill = _WARN_FILL
            if j == 5:
                cell.number_format = _MONEY
                cell.alignment = _CENTER
    _widths(wq, [7, 18, 42, 10, 16, 12, 26])

    # ---------- شیتِ همهٔ موجودی‌های سایت + فلگِ تغییر ----------
    chg = {}
    for c in plan["price_exact"]:
        chg[c["id"]] = ("قیمتِ دقیق", c.get("new"))
    for c in plan["price_family"]:
        chg[c["id"]] = ("قیمتِ خانواده", c.get("new"))
    for c in plan["set_instock"]:
        chg.setdefault(c["id"], ("→ موجود", None))
    for c in plan["set_outofstock"]:
        chg.setdefault(c["id"], ("→ ناموجود", None))
    instock = [p for p in (products or []) if p.get("stock_status") == "instock"]
    ws5 = wb.create_sheet("موجودی‌های سایت")
    ws5.sheet_view.rightToLeft = True
    _header(ws5, ["ردیف", "رفرنس", "نام محصول", "شناسه", "قیمتِ فعلی", "خانواده", "تعدادی", "تغییر", "قیمتِ جدید"])
    for i, p in enumerate(instock, 1):
        change, newp = chg.get(p.get("id"), ("بدونِ تغییر", None))
        qty = "بله" if (p.get("manage_stock") and (p.get("stock_quantity") or 0) >= 1) else "خیر"
        vals = [i, p.get("ref"), (p.get("name") or "")[:50], p.get("id"),
                int(p.get("regular_price", 0) or 0), (p.get("ref") or "").split(".")[0],
                qty, change, (int(newp) if newp else "")]
        for j, v in enumerate(vals, 1):
            cell = ws5.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
            if j in (5, 9):
                cell.number_format = _MONEY
                cell.alignment = _CENTER
        if change != "بدونِ تغییر":                     # ردیفِ تغییرکرده رنگ بخورد
            fill = _UP_FILL if "ناموجود" in change else _DOWN_FILL
            for j in range(1, 10):
                ws5.cell(row=i + 1, column=j).fill = fill
    _widths(ws5, [7, 18, 44, 10, 16, 10, 8, 16, 16])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path
