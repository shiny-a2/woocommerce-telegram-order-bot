"""irantimer_import_job.py — درجِ اکسلِ اصلاح‌شدهٔ اپراتور روی سایت (از طریقِ اندپوینتِ CRM).

هر ردیفِ اکسل → POST /a2crm/v1/tg/import-product (پیش‌نویس، دیدوپ با رفرنس، بدونِ عکس).
⚠️ عکسِ کاتالوگِ منبع روی سایت استفاده نمی‌شود؛ عکس‌ها را مالک خودش ادیت/آپلود و با /media_images می‌چسباند.
تست: IT_IMPORT_DRYRUN=1 (فقط would_create، چیزی ساخته نمی‌شود).
"""
from __future__ import annotations

import json
import os
import sys
import time
import urllib.request
import uuid

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)
os.chdir(_HERE)

import config as c            # noqa: E402
import requests               # noqa: E402
from openpyxl import Workbook, load_workbook          # noqa: E402
from openpyxl.styles import Font, PatternFill          # noqa: E402

_DATA = os.path.join(_HERE, "data")

# نامِ ستونِ اکسل (اتریبیوتِ سایت) → اسلاگِ اتریبیوتِ سراسری
ATTR_SLUG = {
    "مناسب برای": "pa_مناسب-برای", "استایل": "pa_استایل", "طرح صفحه": "pa_طرح-صفحه",
    "رنگ صفحه": "pa_رنگ-صفحه", "شکل قاب": "pa_شکل-قاب", "میزان ضدآبی": "pa_میزان-ضدآبی",
    "جنس بکارگرفته": "pa_جنس-بکارگرفته", "امکانات دیگر": "pa_امکانات-دیگر", "رنگ بند": "pa_رنگ-بند",
    "رنگ قاب": "pa_رنگ-قاب", "نوع موتور": "pa_نوع-موتور", "نوع شیشه": "pa_نوع-شیشه",
    "طرح بند": "pa_طرح-بند", "نوع قفل": "pa_نوع-قفل", "تقویم و نوع آن": "pa_تقویم-و-نوع-آن",
    "نگین": "pa_نگین", "گارانتی": "pa_گارانتی", "گارانتی کننده در ایران": "pa_گارانتی-کننده",
    "موارد گارانتی": "pa_موارد-گارانتی", "اصالت برند": "pa_اصالت-برند", "کشور سازنده": "pa_کشور-سازنده",
    "سایز قاب": "pa_سایز-قاب", "ارتفاع قاب": "pa_ارتفاع-قاب", "عرض بند": "pa_عرض-بند", "وزن ساعت": "pa_وزن-ساعت",
}


def _tg(method, fields, files=None):
    token = c.TELEGRAM_BOT_TOKEN
    if not token:
        return False
    b = uuid.uuid4().hex
    parts = []
    for k, v in fields.items():
        parts.append(f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
    for k, (fn, data, ct) in (files or {}).items():
        parts.append((f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"; filename=\"{fn}\"\r\n"
                      f"Content-Type: {ct}\r\n\r\n").encode())
        parts.append(data)
        parts.append(b"\r\n")
    parts.append(f"--{b}--\r\n".encode())
    req = urllib.request.Request(f"https://api.telegram.org/bot{token}/{method}", data=b"".join(parts),
                                 headers={"Content-Type": f"multipart/form-data; boundary={b}"})
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            return "\"ok\":true" in r.read().decode("utf-8", "replace")
    except Exception:  # noqa: BLE001
        return False


def _recipients():
    ids = list(c.ADMIN_USER_IDS or [])
    op = getattr(c, "WT_MEDIAIMG_OPERATOR_ID", 0)
    if op and op not in ids:
        ids.append(op)
    return ids


def _call_endpoint(ref, name, brand, attrs, price=None, dry=False):
    base = c.CRM_TG_URL
    tok = c.CRM_TG_TOKEN
    if not (base and tok):
        return {"error": "crm_not_configured"}
    params = {"ref": ref, "name": name, "brand": brand, "status": "draft",
              "attrs": json.dumps(attrs, ensure_ascii=False)}
    if price:
        params["regular_price"] = str(price)
    if dry:
        params["dry_run"] = "1"
    # ⚠️ image_url عمداً فرستاده نمی‌شود (عکس‌ها را مالک جدا می‌چسباند)
    try:
        r = requests.post(f"{base}/import-product", params=params,
                          headers={"X-A2-Token": tok, "Accept": "application/json"}, timeout=45)
        r.raise_for_status()
        return r.json()
    except Exception as e:  # noqa: BLE001
        return {"error": type(e).__name__}


def _mk_name(brand, gender, ref):
    g = (gender or "").split(" | ")[0].strip()
    return " ".join(x for x in ["ساعت", g, brand, ref] if x).strip()


def run(xlsx_path, dry=False):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    H = [c0.value for c0 in ws[1]]
    idx = {h: i for i, h in enumerate(H)}
    if "رفرانس" not in idx or "نام برند" not in idx:
        for oid in _recipients():
            _tg("sendMessage", {"chat_id": str(oid), "text": "❌ اکسل ستون‌های «رفرانس»/«نام برند» را ندارد."})
        return 1
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref = str(row[idx["رفرانس"]] or "").strip()
        brand = str(row[idx["نام برند"]] or "").strip()
        if not ref:
            continue
        gender = str(row[idx.get("مناسب برای", -1)] or "") if "مناسب برای" in idx else ""
        attrs = {}
        for col, slug in ATTR_SLUG.items():
            if col in idx:
                val = row[idx[col]]
                if val is not None and str(val).strip():
                    attrs[slug] = [v.strip() for v in str(val).split("|") if v.strip()]
        res = _call_endpoint(ref, _mk_name(brand, gender, ref), brand, attrs, dry=dry)
        res["_ref"] = ref
        results.append(res)
        time.sleep(0.15)
    # گزارش
    created = sum(1 for r in results if r.get("created"))
    skipped = sum(1 for r in results if r.get("skipped"))
    wouldc = sum(1 for r in results if r.get("would_create"))
    errs = sum(1 for r in results if r.get("error"))
    warns = sum(len(r.get("warnings") or []) for r in results)
    rep = os.path.join(_DATA, f"import-{'dry' if dry else 'done'}-{uuid.uuid4().hex[:6]}.xlsx")
    out = Workbook(); s = out.active; s.title = "گزارشِ درج"; s.sheet_view.rightToLeft = True
    s.append(["ردیف", "رفرانس", "نتیجه", "شناسه", "ویژگی‌ها", "هشدارها"])
    for cc in s[1]:
        cc.fill = PatternFill("solid", fgColor="1F4E78"); cc.font = Font(bold=True, color="FFFFFF")
    for i, r in enumerate(results, 1):
        if r.get("created"):
            st = "✅ ساخته شد (پیش‌نویس)"
        elif r.get("would_create"):
            st = "🔸 آماده (dry)"
        elif r.get("skipped"):
            st = "⏭️ تکراری"
        elif r.get("error"):
            st = "❌ " + str(r.get("error"))
        else:
            st = "?"
        s.append([i, r.get("_ref"), st, r.get("product_id", ""),
                  r.get("attrs_set", r.get("attrs_would_set", "")), " ؛ ".join(r.get("warnings") or [])])
    out.save(rep)
    with open(rep, "rb") as f:
        data = f.read()
    head = ("🔸 پیش‌نمایشِ درج (dry — چیزی ساخته نشد)" if dry else "⬆️ درجِ روی سایت انجام شد (پیش‌نویس)")
    cap = (f"{head}\n\n• کلِ ردیف: {len(results)}\n"
           + (f"• آمادهٔ ساخت: {wouldc}\n" if dry else f"• ساخته‌شد (پیش‌نویس): {created}\n")
           + f"• تکراری (رد): {skipped}\n• خطا: {errs}\n• هشدارِ ترمِ نامعتبر: {warns}\n\n"
           + ("برای درجِ واقعی، دکمهٔ «درج» را بزن." if dry else "پیش‌نویس‌ها را ببین و منتشر کن. عکس‌ها را با /media_images بچسبان."))
    for oid in _recipients():
        _tg("sendDocument", {"chat_id": str(oid), "caption": cap},
            {"document": (os.path.basename(rep), data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    return 0


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("IT_IMPORT_FILE", "")
    dry = os.environ.get("IT_IMPORT_DRYRUN", "0") == "1"
    if not path or not os.path.exists(path):
        print("فایل نیست:", path); sys.exit(1)
    sys.exit(run(path, dry=dry))
