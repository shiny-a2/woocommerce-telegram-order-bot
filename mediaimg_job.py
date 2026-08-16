"""mediaimg_job.py — درجِ خودکارِ تصاویرِ محصولاتِ بی‌عکس از کتابخانهٔ رسانه (بر اساسِ رفرنس).

اجرا: از دکمهٔ رباتِ «🖼 درجِ تصاویر» (ادمین‌ها + اپراتور) یا دستی. dry-run پیش‌فرض؛ اعمال با WT_MEDIAIMG_APPLY=1.
گزارشِ اکسل به مالک + اپراتور. نوشتنِ متادیتای کامل (شاملِ description) نیازمندِ اندپوینتِ CRM است
(docs/crm-media-images-endpoint-spec.md)؛ تا آن‌موقع fallback به wc/v3 (شاخص+گالری+title+alt).
"""
from __future__ import annotations

import datetime
import os
import sys
import uuid
import urllib.request

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)
os.chdir(_HERE)

import db          # noqa: E402
db.init()
import config as c          # noqa: E402
import mediaimg as mi       # noqa: E402
from openpyxl import Workbook          # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment          # noqa: E402

_TEHRAN = datetime.timezone(datetime.timedelta(hours=3, minutes=30))
_DATA = os.path.join(_HERE, "data")


def log(msg: str):
    ts = datetime.datetime.now(_TEHRAN).strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")


def _tg(method, fields, files=None):
    token = c.TELEGRAM_BOT_TOKEN
    if not token:
        return False
    b = uuid.uuid4().hex
    parts = []
    for k, v in fields.items():
        parts.append(f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
    for k, (fn, data, ct) in (files or {}).items():
        parts.append((f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"; filename=\"{fn}\"\r\n"
                      f"Content-Type: {ct}\r\n\r\n").encode())
        parts.append(data)
        parts.append(b"\r\n")
    parts.append(f"--{b}--\r\n".encode())
    req = urllib.request.Request(f"https://api.telegram.org/bot{token}/{method}", data=b"".join(parts),
                                 headers={"Content-Type": f"multipart/form-data; boundary={b}"})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return "\"ok\":true" in r.read().decode("utf-8", "replace")
    except Exception:  # noqa: BLE001
        return False


def send_text(chat_id, text):
    return _tg("sendMessage", {"chat_id": str(chat_id), "text": text})


def send_doc(chat_id, path, caption):
    with open(path, "rb") as f:
        data = f.read()
    return _tg("sendDocument", {"chat_id": str(chat_id), "caption": caption},
               {"document": (os.path.basename(path), data,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})


def _recipients():
    ids = list(c.ADMIN_USER_IDS or [])
    op = getattr(c, "WT_MEDIAIMG_OPERATOR_ID", 0)
    if op and op not in ids:
        ids.append(op)
    return ids


_HDR = PatternFill("solid", fgColor="1F4E78")
_HF = Font(bold=True, color="FFFFFF")
_OK = PatternFill("solid", fgColor="E2EFDA")
_BAD = PatternFill("solid", fgColor="FCE4D6")


def build_report(rows, results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "درجِ تصاویر"
    ws.sheet_view.rightToLeft = True
    head = ["ردیف", "شناسه", "نام", "رفرنس", "شاخص", "گالری", "وضعیت", "via", "علت"]
    ws.append(head)
    for cell in ws[1]:
        cell.fill = _HDR
        cell.font = _HF
        cell.alignment = Alignment(horizontal="center")
    for i, (row, res) in enumerate(zip(rows, results), 1):
        status = "✅ اعمال شد" if res.get("done") else ("⛔ بی‌رسانه" if not row["ok"] else "❌ خطا")
        ws.append([i, row["id"], row["name"][:55], row["ref"],
                   row["featured"] or "—", len(row["gallery"]),
                   status, res.get("via") or "—", res.get("err") or row.get("reason") or ""])
        fill = _OK if res.get("done") else _BAD
        for cell in ws[ws.max_row]:
            cell.fill = fill
    widths = [6, 9, 45, 18, 10, 8, 14, 6, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    wb.save(out_path)
    return out_path


def run(apply: bool):
    log(f"اسکنِ {c.WT_MEDIAIMG_SCAN_PAGES} صفحهٔ اخیر برای محصولاتِ بی‌عکس…")
    noimg = mi.fetch_noimage_products(max_pages=c.WT_MEDIAIMG_SCAN_PAGES)
    log(f"بی‌عکس: {len(noimg)} محصول. تطبیقِ رسانه…")
    rows = mi.plan(noimg)
    s = mi.summarize(rows)
    log(f"پلن: {s}")
    results = []
    if apply:
        for row in rows:
            if row["ok"]:
                res = mi.apply_row(row, use_crm=c.WT_MEDIAIMG_USE_CRM)
            else:
                res = {"done": False, "via": None, "err": row.get("reason")}
            results.append(res)
        done = sum(1 for r in results if r.get("done"))
        via_crm = sum(1 for r in results if r.get("via") == "crm")
        log(f"اعمال شد: {done}/{len(rows)} (crm={via_crm}, wc={done - via_crm})")
    else:
        results = [{"done": False, "via": None, "err": None} for _ in rows]
    ts = datetime.datetime.now(_TEHRAN).strftime("%Y%m%d-%H%M%S")
    out = os.path.join(_DATA, f"mediaimg-{'applied' if apply else 'dryrun'}-{ts}.xlsx")
    build_report(rows, results, out)
    return {"rows": rows, "results": results, "summary": s, "xlsx": out}


def main() -> int:
    apply = c.WT_MEDIAIMG_APPLY
    res = run(apply=apply)
    s = res["summary"]
    if apply:
        done = sum(1 for r in res["results"] if r.get("done"))
        crm = sum(1 for r in res["results"] if r.get("via") == "crm")
        cap = (f"🖼 درجِ تصاویر — اعمال شد\n\n• بی‌عکسِ بررسی‌شده: {s['total']}\n"
               f"• عکس‌دار شد: {done}\n• گالری کامل شد: {s['with_gallery']}\n"
               f"• بدونِ رسانه (خطا): {s['no_media']}\n"
               f"• متادیتای کامل (CRM): {crm}" + ("" if crm else " — اندپوینتِ CRM هنوز نیست؛ فقط شاخص+گالری+title+alt"))
    else:
        cap = (f"🖼 درجِ تصاویر — پیش‌نمایش (بدونِ نوشتن)\n\n• بی‌عکس: {s['total']}\n"
               f"• دارای رسانهٔ متناظر: {s['with_featured']}\n• گالری‌دار: {s['with_gallery']}\n"
               f"• بدونِ رسانه: {s['no_media']}\n\nبرای اعمال، WT_MEDIAIMG_APPLY=1.")
    for oid in _recipients():
        send_doc(oid, res["xlsx"], cap)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:  # noqa: BLE001
        import traceback
        log(f"[fatal] {e!r}\n{traceback.format_exc()}")
        sys.exit(1)
