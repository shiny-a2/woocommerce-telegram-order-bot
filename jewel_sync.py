"""سینکِ jeweltime → جواهریان: آینهٔ «موجود/ناموجود + تعدادِ عددی» برای برندهای مبدأ.

- مبدأ: ووکامرسِ jeweltime روی همان سرورِ cPanel (DB `source_products_db`،
  پیشوند `wp_`). خواندن فقط با SSH→mysql (read-only، SELECT).
- مقصد: جواهریان با WooCommerce API (همان مسیرِ امنِ query-string که citizen/pricesync دارند).
- تطبیق: `pa_رفرانس` (نرمال‌شده؛ jeweltime خط‌تیره، جواهریان گاهی نقطه).
- قواعدِ مالک (۲۰۲۶-۰۸-۲۴):
    • موجود/ناموجود + تعدادِ عددی → آینهٔ خودکار (manage_stock=true + stock_quantity).
    • قیمت: فقط گزارش. اقلامِ موجود که قیمتِ jeweltime÷۱۰۰ ≠ قیمتِ جواهریان → در اکسل گزارش
      می‌شوند، بدونِ نوشتنِ خودکار. قیمتِ اقلامِ ناموجود دست نمی‌خورد.
    • استثنا: هانوا «16-6018-13-007» هرگز دست نخورد.
"""
import os
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

_HERE = os.path.dirname(os.path.abspath(__file__))
_DATA = os.path.join(_HERE, "data")

SSH_HOST = "root@source-server.example"
SSH_KEY = os.path.join(_HERE, ".ssh", "jeweltime_ed25519")
JT_DB = "source_products_db"
JT_PREFIX = "wp_"
# اعمالِ سریعِ محلی روی سرور (بدونِ شبکه/حذفِ بدنه/سربارِ per-request): اسکریپتِ PHP روی خودِ سرور،
# جواهریان را با ووکامرس محلی به‌روز می‌کند (بوت‌استرپِ یک‌باره). به‌جای ۱۹۰۰ نوشتنِ HTTPِ ~۲.۴ثانیه‌ای.
PHP_BIN = "/opt/cpanel/ea-php85/root/usr/bin/php"
SERVER_APPLY = "/home/user/jewel_apply_stock.php"
SERVER_USER = "shop"
PRICE_DIVISOR = 100                 # دیتابیسِ jeweltime قیمت را ریال×۱۰۰ ذخیره می‌کند
REF_ATTR = "رفرانس"                 # نامِ اتریبیوتِ رفرنس در attributesِ محصولِ جواهریان
BRAND_TERMS = {                     # برندهای jeweltime روی جواهریان (attr pa_نام-برند id=103)
    "تروساردی": 21619, "تورنادو": 21405, "ژاک فیلیپ": 20947, "سکتور": 22034,
    "فیلیپو لورتی": 22507, "کوریو": 25732, "لوسین روشا": 21277, "هانوا": 21800, "ولدر": 21913,
}
_EXCLUDE = {"16601813007"}          # هانوا 16-6018-13-007 (نرمال‌شده) — مستثنیٰ به‌خواستِ مالک


def norm_ref(s) -> str:
    s = "".join(str(s or "").upper().split())
    for ch in ".-_/\\":
        s = s.replace(ch, "")
    return s


# ---------- خواندنِ مبدأ (jeweltime) با SSH+SQL ----------
_JT_SQL = f"""SET NAMES utf8mb4;
SELECT
 (SELECT t.name FROM {JT_PREFIX}term_relationships tr
    JOIN {JT_PREFIX}term_taxonomy tt ON tt.term_taxonomy_id=tr.term_taxonomy_id AND tt.taxonomy='pa_رفرانس'
    JOIN {JT_PREFIX}terms t ON t.term_id=tt.term_id
  WHERE tr.object_id=p.ID LIMIT 1) AS ref,
 p.post_title AS title,
 MAX(IF(pm.meta_key='_regular_price',pm.meta_value,NULL)) AS price,
 MAX(IF(pm.meta_key='_stock',pm.meta_value,NULL)) AS qty,
 MAX(IF(pm.meta_key='_stock_status',pm.meta_value,NULL)) AS st
FROM {JT_PREFIX}posts p JOIN {JT_PREFIX}postmeta pm ON pm.post_id=p.ID
WHERE p.post_type='product' AND p.post_status='publish'
GROUP BY p.ID HAVING ref IS NOT NULL;
"""


def fetch_jeweltime() -> dict:
    """محصولاتِ jeweltime → {norm_ref: {ref, title, price, qty, status}}. خواندنِ read-only."""
    cmd = ["ssh", "-i", SSH_KEY, "-o", "BatchMode=yes", "-o", "StrictHostKeyChecking=no",
           "-o", "ConnectTimeout=20", SSH_HOST,
           f"mysql {JT_DB} -N --default-character-set=utf8mb4"]
    r = subprocess.run(cmd, input=_JT_SQL.encode("utf-8"),
                       stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=180)
    if r.returncode != 0:
        raise RuntimeError(f"ssh/mysql rc={r.returncode}: {r.stderr.decode('utf-8','replace')[:200]}")
    out = {}
    for line in r.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        ref, title, price, qty, st = parts[0], parts[1], parts[2], parts[3], parts[4]
        if not ref or ref == "NULL":
            continue
        out[norm_ref(ref)] = {"ref": ref, "title": title, "price": price, "qty": qty, "status": st}
    return out


# ---------- خواندنِ مقصد (جواهریان) با WooCommerce API ----------
async def fetch_shop(woo_mod) -> dict:
    """محصولاتِ ۹ برندِ مبدأ روی جواهریان → {norm_ref: {id, ref, name, price, qty, status, manage}}."""
    fields = "id,name,regular_price,stock_status,stock_quantity,manage_stock,attributes"
    out = {}
    for term_id in BRAND_TERMS.values():
        page = 1
        while True:
            batch = await woo_mod.get("products", {"attribute": "pa_نام-برند", "attribute_term": term_id,
                                                   "per_page": 100, "page": page, "_fields": fields})
            if not batch:
                break
            for p in batch:
                ref = ""
                for a in (p.get("attributes") or []):
                    if (a.get("name") or "").strip() == REF_ATTR and a.get("options"):
                        ref = a["options"][0]
                if not ref:
                    continue
                try:
                    rp = int("".join(c for c in str(p.get("regular_price") or "") if c.isdigit()) or 0)
                except ValueError:
                    rp = 0
                out[norm_ref(ref)] = {"id": p.get("id"), "ref": ref, "name": p.get("name") or "",
                                      "price": rp, "qty": p.get("stock_quantity"),
                                      "status": p.get("stock_status"),
                                      "manage": bool(p.get("manage_stock"))}
            if len(batch) < 100:
                break
            page += 1
    return out


# ---------- پلن (خالص، بدونِ I/O) ----------
def _jt_target(jt_row):
    """وضعیت/تعدادِ هدف از رویِ محصولِ jeweltime (آینهٔ عددی)."""
    st = (jt_row.get("status") or "").strip()
    q = jt_row.get("qty")
    try:
        qn = int(q) if q not in (None, "", "NULL") else 0
    except ValueError:
        qn = 0
    if st == "instock":
        return "instock", max(1, qn)      # موجود ⇒ حداقل ۱
    return "outofstock", 0


def plan_changes(jt: dict, jav: dict) -> dict:
    out = {"stock": [], "price_diff": [], "excluded": [], "jt_not_jav": [],
           "jav_not_jt": [], "unchanged": 0}
    for nref, jr in jav.items():
        if nref in _EXCLUDE:
            out["excluded"].append({"ref": jr["ref"], "name": jr["name"], "id": jr["id"]})
            continue
        src = jt.get(nref)
        if not src:
            out["jav_not_jt"].append({"id": jr["id"], "ref": jr["ref"], "name": jr["name"],
                                      "status": jr["status"]})
            continue
        tgt_status, tgt_qty = _jt_target(src)
        cur_status = jr["status"]
        cur_qty = jr["qty"] if jr["qty"] is not None else None
        # نیازِ تغییرِ موجودی: وضعیت فرق دارد، یا تعداد فرق دارد، یا تعدادی نیست
        need = (cur_status != tgt_status) or (not jr["manage"]) or (cur_qty != tgt_qty)
        if need:
            out["stock"].append({"id": jr["id"], "ref": jr["ref"], "name": jr["name"],
                                 "old_status": cur_status, "old_qty": cur_qty,
                                 "new_status": tgt_status, "new_qty": tgt_qty})
        else:
            out["unchanged"] += 1
        # قیمت: فقط گزارش، فقط اقلامِ موجودِ دارای قیمت
        if tgt_status == "instock" and src["price"] not in (None, "", "NULL"):
            try:
                jt_rial = int(src["price"]) // PRICE_DIVISOR
            except ValueError:
                jt_rial = 0
            if jt_rial and jt_rial != jr["price"]:
                out["price_diff"].append({"id": jr["id"], "ref": jr["ref"], "name": jr["name"],
                                          "jav_price": jr["price"], "jt_price": jt_rial})
    jav_refs = set(jav.keys())
    for nref, src in jt.items():
        if nref in _EXCLUDE:
            continue
        if nref not in jav_refs:
            out["jt_not_jav"].append({"ref": src["ref"], "title": src.get("title") or "",
                                      "status": src.get("status"), "qty": src.get("qty")})
    return out


def summarize(plan: dict) -> str:
    return (f"تغییرِ موجودی/تعداد: {len(plan['stock'])} · اختلافِ قیمت(گزارش): {len(plan['price_diff'])} · "
            f"بی‌تغییر: {plan['unchanged']} · مستثنیٰ: {len(plan['excluded'])} · "
            f"jeweltime‌که‌در‌جواهریان‌نیست: {len(plan['jt_not_jav'])} · "
            f"جواهریان‌که‌در‌jeweltime‌نیست: {len(plan['jav_not_jt'])}")


# ---------- اعمال (فقط موجودی/تعداد؛ قیمت هرگز نوشته نمی‌شود) ----------
def apply_plan_server(plan: dict) -> dict:
    """اعمالِ سریعِ محلی: تغییراتِ (id, qty) را به اسکریپتِ PHP روی سرور می‌فرستد (stdin)،
    که با ووکامرسِ محلی manage_stock=true + stock_quantity را ست می‌کند. خروجی OK/ERR per خط."""
    res = {"stock": 0, "errors": []}
    changes = plan["stock"]
    if not changes:
        return res
    lines = "".join(f"{c['id']}\t{c['new_qty']}\n" for c in changes)
    remote = f"su -s /bin/bash {SERVER_USER} -c '{PHP_BIN} {SERVER_APPLY}'"
    cmd = ["ssh", "-i", SSH_KEY, "-o", "BatchMode=yes", "-o", "StrictHostKeyChecking=no",
           "-o", "ConnectTimeout=20", SSH_HOST, remote]
    r = subprocess.run(cmd, input=lines.encode("utf-8"),
                       stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=3600)
    for line in r.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split("\t")
        if parts and parts[0] == "OK":
            res["stock"] += 1
        elif parts and parts[0] == "ERR":
            res["errors"].append({"id": parts[1] if len(parts) > 1 else "?",
                                  "err": parts[2] if len(parts) > 2 else "err"})
    if r.returncode != 0 and res["stock"] == 0:
        res["errors"].append({"id": "-", "err": f"ssh rc={r.returncode}: {r.stderr.decode('utf-8','replace')[:160]}"})
    return res


async def apply_plan(woo_mod, plan: dict, limit=None) -> dict:
    """(fallback) اعمال از راهِ WooCommerce API — کُند (per-request HTTP). مسیرِ اصلی apply_plan_server است."""
    res = {"stock": 0, "errors": []}
    for i, c in enumerate(plan["stock"]):
        if limit and i >= limit:
            break
        try:
            await woo_mod.put(f"products/{c['id']}",
                              {"manage_stock": True, "stock_quantity": c["new_qty"]})
            res["stock"] += 1
        except Exception as e:  # noqa: BLE001
            res["errors"].append({"id": c["id"], "err": type(e).__name__})
    return res


# ---------- گزارشِ اکسل ----------
_HDR = PatternFill("solid", fgColor="1F4E78")
_HF = Font(bold=True, color="FFFFFF")
_CHG = PatternFill("solid", fgColor="E2EFDA")
_OOS = PatternFill("solid", fgColor="FCE4D6")
_PRC = PatternFill("solid", fgColor="FFF2CC")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0"


def _hdr(ws, titles):
    for i, t in enumerate(titles, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = _HDR
        c.font = _HF
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER
    ws.freeze_panes = ws.cell(row=2, column=1)


def _w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(plan: dict, out_path: str, applied=False, apply_result=None) -> str:
    err_ids = {e.get("id") for e in ((apply_result or {}).get("errors") or [])}
    wb = Workbook()
    ws = wb.active
    ws.title = "خلاصه"
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value=f"سینکِ jeweltime → جواهریان — {'اعمال‌شده' if applied else 'پیش‌نمایش'}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="آینهٔ موجود/ناموجود + تعداد. قیمت فقط گزارش می‌شود (jeweltime÷۱۰۰).")
    rows = [("تغییرِ موجودی/تعداد (اعمال)", len(plan["stock"])),
            ("اختلافِ قیمت (فقط گزارش)", len(plan["price_diff"])),
            ("بی‌تغییر", plan["unchanged"]),
            ("مستثنیٰ (هانوا)", len(plan["excluded"])),
            ("در jeweltime هست، جواهریان نیست", len(plan["jt_not_jav"])),
            ("جواهریانِ این برندها که در jeweltime نیست", len(plan["jav_not_jt"]))]
    if applied:
        rows.append(("خطا در نوشتن", len(err_ids)))
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).border = _BORDER
        ws.cell(row=i, column=2, value=v).border = _BORDER
    _w(ws, [42, 12])

    # تغییراتِ موجودی/تعداد
    wsc = wb.create_sheet("موجودی-تعداد")
    wsc.sheet_view.rightToLeft = True
    _hdr(wsc, ["ردیف", "رفرنس", "نام", "شناسه", "وضعیتِ قبلی", "تعدادِ قبلی", "وضعیتِ جدید", "تعدادِ جدید"])
    for i, c in enumerate(plan["stock"], 1):
        vals = [i, c["ref"], (c["name"] or "")[:45], c["id"], c["old_status"],
                c["old_qty"] if c["old_qty"] is not None else "", c["new_status"], c["new_qty"]]
        for j, v in enumerate(vals, 1):
            cell = wsc.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
        fill = _OOS if c["new_status"] == "outofstock" else _CHG
        for j in range(1, 9):
            wsc.cell(row=i + 1, column=j).fill = fill
        if c["id"] in err_ids:
            wsc.cell(row=i + 1, column=2).font = Font(color="C00000", bold=True)
    _w(wsc, [6, 16, 40, 9, 13, 11, 13, 11])

    # اختلافِ قیمت (فقط گزارش)
    wp = wb.create_sheet("اختلاف-قیمت")
    wp.sheet_view.rightToLeft = True
    _hdr(wp, ["ردیف", "رفرنس", "نام", "شناسه", "قیمتِ جواهریان", "قیمتِ jeweltime÷۱۰۰", "اختلاف"])
    for i, c in enumerate(sorted(plan["price_diff"], key=lambda x: -abs((x["jt_price"] or 0) - (x["jav_price"] or 0))), 1):
        diff = (c["jt_price"] or 0) - (c["jav_price"] or 0)
        vals = [i, c["ref"], (c["name"] or "")[:45], c["id"], c["jav_price"], c["jt_price"], diff]
        for j, v in enumerate(vals, 1):
            cell = wp.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
            if j in (5, 6, 7):
                cell.number_format = _MONEY
            cell.fill = _PRC
    _w(wp, [6, 16, 40, 9, 16, 18, 16])

    # در jeweltime هست، جواهریان نیست
    wj = wb.create_sheet("jeweltime‌نبود‌در‌جواهریان")
    wj.sheet_view.rightToLeft = True
    _hdr(wj, ["ردیف", "رفرنس", "نام", "وضعیت", "تعداد"])
    for i, c in enumerate(plan["jt_not_jav"], 1):
        for j, v in enumerate([i, c["ref"], (c["title"] or "")[:45], c.get("status"), c.get("qty")], 1):
            wj.cell(row=i + 1, column=j, value=v).border = _BORDER
    _w(wj, [6, 16, 40, 12, 8])

    # جواهریانِ این برندها که در jeweltime نیست
    wv = wb.create_sheet("جواهریان‌نبود‌در‌jeweltime")
    wv.sheet_view.rightToLeft = True
    _hdr(wv, ["ردیف", "رفرنس", "نام", "شناسه", "وضعیتِ فعلی"])
    for i, c in enumerate(plan["jav_not_jt"], 1):
        for j, v in enumerate([i, c["ref"], (c["name"] or "")[:45], c["id"], c["status"]], 1):
            wv.cell(row=i + 1, column=j, value=v).border = _BORDER
    _w(wv, [6, 16, 40, 9, 13])

    wb.save(out_path)
    return out_path


# ---------- ران ----------
async def run(woo_mod, apply: bool = False, out_dir: str | None = None) -> dict:
    jt = fetch_jeweltime()
    jav = await fetch_shop(woo_mod)
    plan = plan_changes(jt, jav)
    result = apply_plan_server(plan) if apply else None   # اعمالِ سریعِ محلی روی سرور
    out_dir = out_dir or _DATA
    import time
    ts = time.strftime("%Y%m%d-%H%M%S")
    out = os.path.join(out_dir, f"jewel-{'applied' if apply else 'dryrun'}-{ts}.xlsx")
    build_report(plan, out, applied=apply, apply_result=result)
    return {"plan": plan, "jt_count": len(jt), "jav_count": len(jav),
            "result": result, "xlsx": out, "summary": summarize(plan)}
