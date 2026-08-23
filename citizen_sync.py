"""citizen_sync.py — سینکِ قیمت/موجودیِ برندِ «سیتیزن» از تأمین‌کنندهٔ app.supplier.example به ووکامرس.

منبع = APIِ سیتیزن (saati.py): reference + price(ریال) + count(موجودی). تطبیق با اتریبیوتِ «رفرانس»ِ سایت (دقیق).
قواعد (تأییدِ مالک): قیمت مستقیم (ریال، بدونِ تغییرِ واحد)؛ موجودی = فقط وضعیت (count≥۱→instock، ۰→outofstock)؛
محصولِ سیتیزنِ سایت که در تأمین‌کننده نیست → outofstock. موجودیِ تعدادی اگر بود، مقدارش دست‌نخورده (فقط وضعیت/قیمت).

نوشتن با woo.put (query-string؛ همان مسیرِ اثبات‌شدهٔ pricesync — میزبان بدنه را strip می‌کند).
گزارشِ کاملِ اکسل: خلاصه + تغییرات + «همهٔ رفرنس‌های سایت» (داریم/نداریم/موجودشده/قیمت‌تغییر/بی‌قیمت) + تأمین‌که‌در‌سایت‌نیست.
"""
from __future__ import annotations

import os
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import saati
import woo

CITIZEN_BRAND_TERM = 3197          # ترمِ اتریبیوتِ pa_نام-برند برای «سیتیزن»
QQ_BRAND_TERM = 28861              # ترمِ برندِ «کیو اند کیو» (Q&Q)
# تأمین‌کنندهٔ ساعتی چندبرندی است: category1 «QQ*» = Q&Q، بقیه (Mechanical/EcoDrive/Quartz/Promaster/…) = سیتیزن.
BRAND_CFG = {
    "citizen": {"site_term": CITIZEN_BRAND_TERM, "is_qq": False, "label": "سیتیزن"},
    "qq":      {"site_term": QQ_BRAND_TERM,       "is_qq": True,  "label": "کیو اند کیو"},
}
REF_ATTR = "رفرانس"
BRAND_ATTR = "نام برند"
_DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")


def norm_ref(r) -> str:
    """رفرنس را یکدست می‌کند: بزرگ، بدونِ فاصله. «at2568-82e»→«AT2568-82E»."""
    return "".join(str(r or "").upper().split())


# ---------- واکشی ----------
def fetch_supplier(brand: str = "citizen") -> tuple[dict, dict]:
    """{ref: {price:int(ریال), count:int, name}} فقط برای برندِ خواسته‌شده + آمار. unauthorized → استثنا."""
    want_qq = BRAND_CFG[brand]["is_qq"]
    res = saati.fetch_all_products()
    if res.get("unauthorized"):
        raise PermissionError("saati_unauthorized")
    out = {}
    skipped_other = 0
    for it in res.get("items") or []:
        is_qq = str(it.get("category1") or "").upper().startswith("QQ")
        if is_qq != want_qq:      # فقط محصولاتِ همین برند (QQ*=Q&Q، بقیه=سیتیزن) — لیست‌ها جدا می‌مانند
            skipped_other += 1
            continue
        ref = norm_ref(it.get("reference"))
        if not ref:
            continue
        out[ref] = {"price": int(it.get("price") or 0), "count": int(it.get("count") or 0),
                    "name": (it.get("name") or "").strip()}
    return out, {"total": len(res.get("items") or []), "pages": res.get("pages"), "refs": len(out),
                 "skipped_other": skipped_other, "brand": brand}


async def fetch_site(brand: str = "citizen") -> list:
    """همهٔ محصولاتِ سایتِ برندِ خواسته‌شده. خروجی = list[{id, ref, name, price, status, manage_stock, qty}]."""
    term = BRAND_CFG[brand]["site_term"]
    out, page = [], 1
    fields = "id,name,regular_price,stock_status,manage_stock,stock_quantity,attributes"
    while True:
        batch = await woo.get("products", {"attribute": "pa_نام-برند", "attribute_term": term,
                                           "per_page": 100, "page": page, "_fields": fields})
        if not batch:
            break
        for p in batch:
            ref = ""
            for a in (p.get("attributes") or []):
                if (a.get("name") or "").strip() == REF_ATTR and a.get("options"):
                    ref = a["options"][0]
            try:
                rp = int("".join(c for c in str(p.get("regular_price") or "") if c.isdigit()) or 0)
            except ValueError:
                rp = 0
            out.append({"id": p.get("id"), "ref": norm_ref(ref), "name": p.get("name") or "",
                        "price": rp, "status": p.get("stock_status"),
                        "manage_stock": bool(p.get("manage_stock")), "qty": p.get("stock_quantity")})
        if len(batch) < 100:
            break
        page += 1
    return out


# ---------- پلن ----------
def plan_changes(supplier: dict, site: list) -> dict:
    """محاسبهٔ خالص (بدونِ I/O). قیمت مستقیم؛ موجودی فقط وضعیت؛ سایتِ‌نبود‌در‌تأمین → ناموجود."""
    out = {"price": [], "instock": [], "outofstock": [], "no_price": [], "rows": [],
           "unmatched_site": [], "supplier_not_site": []}
    site_refs = set()
    for s in site:
        ref = s["ref"]
        site_refs.add(ref)
        sup = supplier.get(ref)
        row = {"id": s["id"], "ref": ref, "name": s["name"], "cur_price": s["price"],
               "cur_status": s["status"], "in_supplier": bool(sup),
               "sup_price": sup["price"] if sup else None, "sup_count": sup["count"] if sup else None,
               "change": ""}
        if not ref:
            row["change"] = "بی‌رفرنس"
            out["rows"].append(row)
            continue
        if sup:
            changes = []
            in_stock = sup["count"] >= 1
            # قیمت فقط وقتی تأمین‌کننده «موجود» است اعمال می‌شود. قیمتِ محصولِ ناموجودِ ساعتی
            # کهنه/غیرقابل‌اعتماد است و نباید قیمتِ (دستیِ) سایتِ ما را دست بزند/کاهش دهد.
            if in_stock:
                if sup["price"] > 0 and s["price"] != sup["price"]:
                    out["price"].append({"id": s["id"], "ref": ref, "old": s["price"], "new": sup["price"]})
                    changes.append("قیمت")
                elif sup["price"] == 0:
                    out["no_price"].append({"id": s["id"], "ref": ref})
                    changes.append("قیمتِ تأمین‌کننده صفر")
            elif sup["price"] > 0 and s["price"] != sup["price"]:
                changes.append("قیمتِ ناموجود—دست‌نخورد")  # فقط گزارش؛ نوشته نمی‌شود
            want = "instock" if in_stock else "outofstock"
            if s["status"] != want:
                (out["instock"] if want == "instock" else out["outofstock"]).append({"id": s["id"], "ref": ref})
                changes.append("→موجود" if want == "instock" else "→ناموجود")
            row["change"] = " + ".join(changes) if changes else "بدونِ تغییر"
        else:
            out["unmatched_site"].append(row)
            if s["status"] != "outofstock":
                out["outofstock"].append({"id": s["id"], "ref": ref})
                row["change"] = "در تأمین‌کننده نیست → ناموجود"
            else:
                row["change"] = "در تأمین‌کننده نیست"
        out["rows"].append(row)
    out["supplier_not_site"] = [{"ref": r, "price": v["price"], "count": v["count"], "name": v["name"]}
                                for r, v in supplier.items() if r not in site_refs]
    return out


def summarize(plan: dict) -> str:
    return (f"قیمت: {len(plan['price'])} · →موجود: {len(plan['instock'])} · →ناموجود: {len(plan['outofstock'])} · "
            f"بی‌قیمتِ تأمین: {len(plan['no_price'])} · سایتِ‌نبود‌در‌تأمین: {len(plan['unmatched_site'])} · "
            f"تأمینِ‌نبود‌در‌سایت: {len(plan['supplier_not_site'])} · کلِ ردیف: {len(plan['rows'])}")


async def apply_plan(woo_mod, plan: dict, limit=None) -> dict:
    """نوشتنِ واقعی (فقط فیلدِ لازم). قیمت رشته؛ موجودی stock_status."""
    res = {"price": 0, "instock": 0, "outofstock": 0, "errors": []}
    n = 0

    async def _put(pid, payload, kind):
        nonlocal n
        try:
            await woo_mod.put(f"products/{pid}", payload)
            res[kind] += 1
        except Exception as e:  # noqa: BLE001
            res["errors"].append({"id": pid, "kind": kind, "err": type(e).__name__})
        n += 1

    for c in plan["price"]:
        if limit and n >= limit:
            return res
        await _put(c["id"], {"regular_price": str(c["new"])}, "price")
    for c in plan["instock"]:
        if limit and n >= limit:
            return res
        await _put(c["id"], {"stock_status": "instock"}, "instock")
    for c in plan["outofstock"]:
        if limit and n >= limit:
            return res
        await _put(c["id"], {"stock_status": "outofstock"}, "outofstock")
    return res


# ---------- گزارشِ اکسل ----------
_HDR = PatternFill("solid", fgColor="1F4E78")
_HF = Font(bold=True, color="FFFFFF")
_WARN = PatternFill("solid", fgColor="FFF2CC")
_CHG = PatternFill("solid", fgColor="E2EFDA")
_OOS = PatternFill("solid", fgColor="FCE4D6")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0"


def _hdr(ws, titles):
    for i, t in enumerate(titles, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.fill = _HDR
        c.font = _HF
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER
    ws.freeze_panes = ws.cell(row=2, column=1)


def _w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(plan: dict, meta: dict, out_path: str, applied=False, apply_result=None) -> str:
    err_ids = {e.get("id") for e in ((apply_result or {}).get("errors") or [])}
    wb = Workbook()
    # خلاصه
    ws = wb.active
    ws.title = "خلاصه"
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value=f"گزارشِ سینکِ سیتیزن — {'اعمال‌شده' if applied else 'پیش‌نمایش'}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"تأمین‌کننده: {meta.get('total')} محصول · {meta.get('refs')} رفرنس · سایت: {len(plan['rows'])} سیتیزن")
    rows = [("تغییرِ قیمت", len(plan["price"])), ("→ موجود", len(plan["instock"])),
            ("→ ناموجود", len(plan["outofstock"])), ("موجود ولی قیمتِ تأمین صفر", len(plan["no_price"])),
            ("سایت که در تأمین‌کننده نیست", len(plan["unmatched_site"])),
            ("تأمین‌کننده که در سایت نیست", len(plan["supplier_not_site"]))]
    if applied:
        rows.append(("خطا در نوشتن", len(err_ids)))
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).border = _BORDER
        ws.cell(row=i, column=2, value=v).border = _BORDER
    _w(ws, [40, 12])

    # همهٔ رفرنس‌های سیتیزنِ سایت (گزارشِ کامل)
    wa = wb.create_sheet("همهٔ رفرنس‌های سایت")
    wa.sheet_view.rightToLeft = True
    _hdr(wa, ["ردیف", "رفرنس", "نام", "شناسه", "قیمتِ سایت", "وضعیتِ سایت",
              "در تأمین‌کننده؟", "قیمتِ تأمین", "count", "نتیجهٔ سینک"])
    for i, r in enumerate(sorted(plan["rows"], key=lambda x: (not x["in_supplier"], x["ref"])), 1):
        vals = [i, r["ref"], (r["name"] or "")[:45], r["id"], r["cur_price"], r["cur_status"],
                "بله" if r["in_supplier"] else "خیر", r["sup_price"] if r["sup_price"] is not None else "",
                r["sup_count"] if r["sup_count"] is not None else "", r["change"]]
        for j, v in enumerate(vals, 1):
            cell = wa.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
            if j in (5, 8):
                cell.number_format = _MONEY
        if r["change"] and r["change"] not in ("بدونِ تغییر", "در تأمین‌کننده نیست"):
            fill = _OOS if "ناموجود" in r["change"] else _CHG
            for j in range(1, 11):
                wa.cell(row=i + 1, column=j).fill = fill
        if r["id"] in err_ids:
            wa.cell(row=i + 1, column=2).font = Font(color="C00000", bold=True)
    _w(wa, [6, 16, 40, 9, 15, 13, 12, 15, 8, 26])

    # تأمین‌کننده که در سایت نیست (برای افزودنِ محصول)
    wn = wb.create_sheet("تأمین‌که‌در‌سایت‌نیست")
    wn.sheet_view.rightToLeft = True
    _hdr(wn, ["ردیف", "رفرنس", "نام", "قیمتِ تأمین", "count"])
    for i, s in enumerate(sorted(plan["supplier_not_site"], key=lambda x: -x["count"]), 1):
        vals = [i, s["ref"], (s["name"] or "")[:50], s["price"], s["count"]]
        for j, v in enumerate(vals, 1):
            cell = wn.cell(row=i + 1, column=j, value=v)
            cell.border = _BORDER
            if j == 4:
                cell.number_format = _MONEY
    _w(wn, [6, 16, 48, 15, 8])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


async def run(woo_mod, apply: bool = False, out_dir: str | None = None, brand: str = "citizen") -> dict:
    supplier, meta = fetch_supplier(brand)
    site = await fetch_site(brand)
    plan = plan_changes(supplier, site)
    result = await apply_plan(woo_mod, plan) if apply else None
    out_dir = out_dir or _DATA
    ts = time.strftime("%Y%m%d-%H%M%S")
    tag = "citizen" if brand == "citizen" else brand
    out = os.path.join(out_dir, f"{tag}-{'applied' if apply else 'dryrun'}-{ts}.xlsx")
    build_report(plan, meta, out, applied=apply, apply_result=result)
    return {"plan": plan, "meta": meta, "result": result, "xlsx": out, "summary": summarize(plan)}
