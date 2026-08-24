"""irantimer_extract_job.py — دکمهٔ «گرفتنِ اکسلِ برند» برای اپراتور.

برند → Brand IDِ کاتالوگِ منبع → محصولاتِ آن برند → فقط آن‌هایی که روی سایت «نداریم» (دیدوپ با رفرنس)
→ نگاشتِ کامل (irantimer_map) → اکسلِ عکس‌دار (۱۵۰×۱۵۰، ستونِ اول) → ارسال به اپراتور + مالک.
بسته‌ای: هر بار `batch` محصول از `offset`. برای ادامه، offset بعدی.
"""
from __future__ import annotations

import io
import json
import os
import sys
import urllib.request
import uuid

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)
os.chdir(_HERE)

import config as c            # noqa: E402
import irantimer as it        # noqa: E402
import irantimer_map as im    # noqa: E402
import woo                    # noqa: E402
from openpyxl import Workbook          # noqa: E402
from openpyxl.drawing.image import Image as XLImage          # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill          # noqa: E402
from openpyxl.utils import get_column_letter          # noqa: E402
from PIL import Image as PILImage          # noqa: E402
import requests               # noqa: E402

_UA = {"User-Agent": "Mozilla/5.0"}
_DATA = os.path.join(_HERE, "data")

# برندِ فارسی → (Brand IDِ کاتالوگِ منبع، ترمِ برندِ سایت). ترمِ سایت پویا هم گرفته می‌شود.
BRANDS = {
    "سیتیزن": (10, 3197), "لی کوپر": (227, 5458), "لی‌کوپر": (227, 5458),
    "دنیل کلین": (275, 4695), "دنیل‌کلین": (275, 4695), "دنیل گورمن": (395, None),
    "کاترپیلار": (163, 5422), "تیسوت": (20, None), "پیر لنون": (205, None), "پیرلنون": (205, None),
    "فسیل": (54, 4812), "امپریو آرمانی": (55, 1199), "ادوکس": (23, 3882), "رومانسون": (170, 5528),
    "سواچ": (153, 1833), "استورم": (36, 14994), "اورینت": (18, 6519), "آیس واچ": (122, 25259),
    "تایم فورس": (78, 7018), "مایکل کورس": (75, 4344), "فری لوک": (248, 4964), "ریباک": (157, 9286),
    "سانتاباربارا": (247, 5110), "پولو سانتا باربارا": (247, 5110),
}


def _tg(method, fields, files=None):
    token = c.TELEGRAM_BOT_TOKEN
    if not token:
        return False
    b = uuid.uuid4().hex
    parts = []
    for k, v in fields.items():
        parts.append(f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
    for k, (fn, data, ct) in (files or {}).items():
        parts.append((f"--{b}\r\nContent-Disposition: form-data; name=\"{k}\"; filename=\"{fn}\"\r\n"
                      f"Content-Type: {ct}\r\n\r\n").encode())
        parts.append(data)
        parts.append(b"\r\n")
    parts.append(f"--{b}--\r\n".encode())
    req = urllib.request.Request(f"https://api.telegram.org/bot{token}/{method}", data=b"".join(parts),
                                 headers={"Content-Type": f"multipart/form-data; boundary={b}"})
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            return "\"ok\":true" in r.read().decode("utf-8", "replace")
    except Exception:  # noqa: BLE001
        return False


def _recipients():
    ids = list(c.ADMIN_USER_IDS or [])
    op = getattr(c, "WT_MEDIAIMG_OPERATOR_ID", 0)
    if op and op not in ids:
        ids.append(op)
    return ids


def _norm(s):
    return (s or "").strip()


def resolve_brand(name):
    name = _norm(name)
    for k, (bid, term) in BRANDS.items():
        if name == k or name.replace(" ", "") == k.replace(" ", ""):
            if term is None:  # ترمِ سایت را پویا پیدا کن
                term = _site_term(name)
            return bid, term, k
    return None, None, None


def _site_term(name):
    try:
        terms = woo._get_sync("products/attributes/103/terms", {"per_page": 100, "search": name, "_fields": "id,name"})
        for t in terms:
            if _norm(t.get("name")) == name:
                return t["id"]
        return terms[0]["id"] if terms else None
    except Exception:  # noqa: BLE001
        return None


def site_refs_for(term_id):
    """همهٔ رفرنس‌های سایت برای یک برند (برای دیدوپ)."""
    refs = set()
    if not term_id:
        return refs
    page = 1
    while page <= 60:
        ps = woo._get_sync("products", {"attribute": "pa_نام-برند", "attribute_term": str(term_id),
                                        "per_page": 100, "page": page, "_fields": "id,attributes"})
        if not ps:
            break
        for p in ps:
            for a in p.get("attributes", []):
                if a.get("name") == "رفرانس":
                    for o in (a.get("options") or []):
                        refs.add(o.upper().replace(" ", ""))
        if len(ps) < 100:
            break
        page += 1
    return refs


def _build_excel(rows_details, brand, path):
    """rows_details: list of (mapped_row_dict, detail_dict). اکسلِ عکس‌دار می‌سازد."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{brand} (جدید)"
    ws.sheet_view.rightToLeft = True
    ref_cols = ["نامِ‌کاتالوگِ منبع(مرجع)", "قیمتِ‌تومان(مرجع)", "موجودی", "لینک"]
    headers = ["عکس"] + im.OUT_ATTRS + ref_cols
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    r_i = 2
    imgs = 0
    for row, d in rows_details:
        ws.cell(row=r_i, column=1, value="")
        for j, h in enumerate(im.OUT_ATTRS):
            ws.cell(row=r_i, column=j + 2, value=row.get(h, ""))
        base = 2 + len(im.OUT_ATTRS)
        ws.cell(row=r_i, column=base + 0, value=d.get("title", ""))
        ws.cell(row=r_i, column=base + 1, value=d.get("price_toman"))
        ws.cell(row=r_i, column=base + 2, value="موجود" if d.get("in_stock") else "ناموجود")
        ws.cell(row=r_i, column=base + 3, value=d.get("url", ""))
        ws.row_dimensions[r_i].height = 115
        url = d.get("image")
        if url:
            try:
                b = requests.get(url, headers=_UA, timeout=25).content
                pim = PILImage.open(io.BytesIO(b)).convert("RGB").resize((150, 150))
                buf = io.BytesIO(); pim.save(buf, format="PNG"); buf.seek(0)
                xi = XLImage(buf); xi.width = 150; xi.height = 150
                ws.add_image(xi, f"A{r_i}")
                imgs += 1
            except Exception:  # noqa: BLE001
                pass
        r_i += 1
    wb.save(path)
    return imgs


def run(brand_name, offset=0, batch=None):
    bid, term, canon = resolve_brand(brand_name)
    if not bid:
        for oid in _recipients():
            _tg("sendMessage", {"chat_id": str(oid),
                "text": f"❌ برندِ «{brand_name}» در نگاشتِ کاتالوگِ منبع نیست. برندهای موجود: {'، '.join(sorted(set(BRANDS)))}"})
        return 1
    for oid in _recipients():
        _tg("sendMessage", {"chat_id": str(oid),
            "text": f"⏳ در حالِ استخراجِ کاملِ «{canon}» از کاتالوگِ منبع… (کل برند، ممکن است چند دقیقه طول بکشد)"})
    prods = it.list_products(bid, 1)
    have = site_refs_for(term)
    new = [p for p in prods if p.get("ref") and p["ref"].upper().replace(" ", "") not in have]
    total_new = len(new)
    # پیش‌فرض: کلِ برند در یک فایل (batch=None). offset فقط برای ادامهٔ دستی.
    chunk = new[offset:] if batch is None else new[offset:offset + batch]
    rows_details = []
    for p in chunk:
        try:
            d = it.parse_detail(str(p["id"]))
            rows_details.append((im.map_product(d, canon), d))
        except Exception:  # noqa: BLE001
            continue
    ts_path = os.path.join(_DATA, f"irantimer-{canon}-{offset}-{offset+len(chunk)}.xlsx")
    imgs = _build_excel(rows_details, canon, ts_path)
    with open(ts_path, "rb") as f:
        data = f.read()
    if batch is None:
        span, more = f"کلِ برند: {len(rows_details)} محصول", "\n\n✅ کاملِ این برند استخراج شد."
    else:
        nxt = offset + batch
        span = f"این بسته: {offset+1} تا {offset+len(chunk)} ({len(rows_details)} محصول)"
        more = f"\n\nبرای {min(batch, total_new - nxt)} محصولِ بعدی، دوباره درخواست بده (از {nxt})." if nxt < total_new else "\n\nاین آخرین بسته بود."
    cap = (f"📥 «{canon}» — محصولاتِ جدید (که روی سایت نداریم)\n\n"
           f"• کلِ جدید: {total_new}\n• {span}، {imgs} عکس\n\n"
           f"قوانینت اعمال شده. با عکس تطبیق بده و اصلاحات را ریپلای کن.{more}")
    ok = True
    for oid in _recipients():
        ok = _tg("sendDocument", {"chat_id": str(oid), "caption": cap},
                 {"document": (os.path.basename(ts_path), data,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}) and ok
    return 0


if __name__ == "__main__":
    brand = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("IT_BRAND", "")
    off = int(sys.argv[2]) if len(sys.argv) > 2 else int(os.environ.get("IT_OFFSET", "0"))
    sys.exit(run(brand, offset=off))
