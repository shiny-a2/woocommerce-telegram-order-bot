"""irantimer_map.py — نگاشتِ کاملِ specsِ کاتالوگِ منبع → فرمتِ اتریبیوتِ سایت.

ترکیبِ:
  • قوانینِ قطعیِ اپراتور (data/atefeh_rules.md): طرح صفحه، استایل، زمان‌سنج→کورنوگراف، رنگ‌ها، جداکنندهٔ « | ».
  • نگاشت‌های یادگرفته‌شدهٔ دفترچه (data/citizen_daftarche.xlsx) برای بقیهٔ ویژگی‌ها (برندمستقل، فیزیکی).

خروجی برای دکمهٔ «گرفتنِ اکسلِ برند» و (بعداً) درجِ روی سایت استفاده می‌شود. برندمستقل است چون
ویژگی‌های فیزیکی (رنگ/جنس/موتور/شیشه/قفل/شکل) بینِ برندها مشترک‌اند؛ ثابت‌های برند خالی می‌مانند تا اپراتور پر کند.
"""
from __future__ import annotations

import os
import re
from collections import OrderedDict

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
_DAFTAR = os.path.join(_HERE, "data", "citizen_daftarche.xlsx")

_SEP = " | "  # جداکنندهٔ چندمقداریِ سایت (طبقِ اپراتور)
_COLOR_MAP = {"نقره ای": "سیلور", "نقره‌ای": "سیلور", "رز گلد": "رزگلد", "رزگولد": "رزگلد", "نوک مدادی": "خاکستری"}


def _norm(s):
    return re.sub(r"\s+", " ", (s or "").strip())


# ---------- بارگذاریِ نگاشت‌های ولیوی دفترچه ----------
def _load_value_maps():
    maps = {}
    try:
        wb = openpyxl.load_workbook(_DAFTAR)
    except Exception:  # noqa: BLE001
        return maps
    for sn in wb.sheetnames:
        if sn in ("خلاصه", "ثابت‌ها", "بدون‌نگاشت"):
            continue
        m = {}
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            site_term = r[0] if r else None
            iran = r[1] if len(r) > 1 else None
            if not site_term or not iran or str(iran).strip() in ("—", "-", ""):
                continue
            for part in str(iran).split("؛"):
                p = _norm(part)
                if p:
                    m[p] = site_term
        if m:
            maps[sn] = m
    return maps


VALUE_MAPS = _load_value_maps()

# نامِ اتریبیوتِ سایت ← کلیدِ specsِ کاتالوگِ منبع (برای ویژگی‌های value-map)
SRC = {
    "مناسب برای": "جنسیت", "رنگ صفحه": "رنگ صفحه", "شکل قاب": "شکل قاب",
    "میزان ضدآبی": "مقاوم در برابر آب", "جنس بکارگرفته": "جنس بکاررفته",
    "نوع موتور": "نوع موتور", "نوع شیشه": "جنس شیشه", "طرح بند": "طرح بند",
    "نوع قفل": "نوع قفل", "تقویم و نوع آن": "تقویم", "اصالت برند": "اصالت کشور برند",
    "کشور سازنده": "اصالت کشور برند", "موارد گارانتی": "موارد گارانتی",
}

# ترتیبِ ستون‌های خروجی (فرمتِ سایت؛ «رنگ بکاررفته» حذف طبقِ اپراتور)
OUT_ATTRS = [
    "نام برند", "رفرانس", "مناسب برای", "استایل", "طرح صفحه", "رنگ صفحه", "شکل قاب",
    "میزان ضدآبی", "جنس بکارگرفته", "امکانات دیگر", "رنگ بند", "رنگ قاب", "نوع موتور",
    "نوع شیشه", "طرح بند", "نوع قفل", "تقویم و نوع آن", "نگین", "گارانتی",
    "گارانتی کننده در ایران", "موارد گارانتی", "اصالت برند", "کشور سازنده",
    "سایز قاب", "ارتفاع قاب", "عرض بند", "وزن ساعت",
]


# ---------- قوانینِ اپراتور ----------
def _has_chrono(features):
    return any(t in (features or "") for t in ("کورنوگراف", "کرنوگراف", "زمان سنج", "زمان‌سنج", "کرونوگراف"))


def rule_tarh(specs):
    ts = specs.get("طرح صفحه", "") or ""
    if "چند عقربه" in ts:
        return "آنالوگ (چند عقربه - چند موتوره)" if _has_chrono(specs.get("ویژگی", "")) else "آنالوگ (چند عقربه - تک موتوره)"
    if "دیجیتال" in ts:
        return "دیجیتال"
    if "عقربه" in ts or "آنالوگ" in ts:
        return "آنالوگ (تک عقربه - تک موتوره)"
    return ""


def rule_style(specs):
    ts = specs.get("طرح صفحه", "") or ""
    if "چند عقربه" in ts:
        return "اسپرت - کلاسیک"
    if "عقربه" in ts or "آنالوگ" in ts:
        return "کلاسیک"
    return ""


# نگاشتِ ویژگیِ کاتالوگِ منبع → ترمِ دقیقِ «امکانات دیگر»‌یِ سایت (اسکنِ زیررشته‌ای؛ هرچه تریگر ندارد حذف).
# طبقِ اپراتور: فقط ترم‌های موجودِ سایت مجازند؛ تقویم/اکودرایو/شاخص‌ذخیره/صرفه‌جویی حذف (تریگر ندارند).
_FEATURE_MAP = [
    ("شب نما", "عقربه شب نما"),
    ("زمان سنج", "کورنوگراف"), ("زمان‌سنج", "کورنوگراف"), ("کورنوگراف", "کورنوگراف"),
    ("کرنوگراف", "کورنوگراف"), ("کرونوگراف", "کورنوگراف"),
    ("سرعت سنج", "تاچیمتر"), ("تاچیمتر", "تاچیمتر"),
    ("ساعت جهانی", "نشانگر ساعت جهانی"),
    ("دو زمانه", "دو زمانه"), ("دوزمانه", "دو زمانه"),
    ("زنگ هشدار", "زنگ هشدار"), ("آلارم", "زنگ هشدار"),
    ("کرنومتر", "کرنومتر"),
    ("درب پشت شیشه", "درب پشت شیشه ای"),
    ("زیرثانیه", "زیرثانیه"), ("زیر ثانیه", "زیرثانیه"),
    ("شمارش معکوس", "تایمر شمارش معکوس"),
    ("حالت ماه", "حالت ماه"), ("فاز ماه", "حالت ماه"),
    ("اتصال به گوشی", "قابلیت اتصال به گوشی"), ("بلوتوث", "قابلیت اتصال به گوشی"),
    ("اسکلتون", "اپن هارت - اسکلتون"), ("اپن هارت", "اپن هارت - اسکلتون"),
    ("نور پشت صفحه", "نور پشت صفحه"),
    ("زه قاب چرخشی", "زه قاب چرخشی"), ("بازل چرخان", "زه قاب چرخشی"),
]


def rule_features(specs):
    blob = specs.get("ویژگی", "") or ""
    out = []
    for trig, term in _FEATURE_MAP:
        if trig in blob and term not in out:
            out.append(term)
    return _SEP.join(out)


def norm_color(v):
    v = _norm(v)
    if not v or "ترکیب چند رنگ" in v:
        return ""
    out = []
    for p in re.split(r"[/،,]", v):
        p = _norm(p)
        if not p:
            continue
        p = _COLOR_MAP.get(p, p)
        if p not in out:
            out.append(p)
    return _SEP.join(out)


def _mapv(site_attr, specs):
    """نگاشتِ value-map برای یک ویژگی از specs (با «؛» چندمقداری)."""
    src = SRC.get(site_attr)
    raw = _norm(specs.get(src, "")) if src else ""
    if not raw:
        return ""
    vmap = VALUE_MAPS.get(site_attr, {})
    if not vmap:
        return raw  # نگاشتی نیست → خام (اپراتور بازبینی می‌کند)
    parts = []
    for p in re.split(r"[/،]", raw):
        p = _norm(p)
        if not p:
            continue
        term = vmap.get(p, "")
        if term and term not in parts:
            parts.append(term)
    return _SEP.join(parts) if parts else ("" )  # ناموجود در نگاشت → خالی (بازبینیِ اپراتور)


def _size(specs, key, unit="mm"):
    v = _norm(specs.get(key, ""))
    m = re.search(r"[\d.]+", v)
    return f"{m.group(0)}{unit}" if m else ""


def map_product(d: dict, brand: str) -> "OrderedDict":
    """یک محصولِ کاتالوگِ منبع (خروجیِ irantimer.parse_detail) → dictِ اتریبیوتِ سایت."""
    sp = d.get("specs", {}) or {}
    row = OrderedDict()
    row["نام برند"] = brand
    row["رفرانس"] = d.get("ref") or ""
    row["مناسب برای"] = _mapv("مناسب برای", sp)
    row["استایل"] = rule_style(sp)
    row["طرح صفحه"] = rule_tarh(sp)
    row["رنگ صفحه"] = norm_color(sp.get("رنگ صفحه", ""))
    row["شکل قاب"] = _mapv("شکل قاب", sp)
    row["میزان ضدآبی"] = _mapv("میزان ضدآبی", sp)
    row["جنس بکارگرفته"] = _mapv("جنس بکارگرفته", sp)
    row["امکانات دیگر"] = rule_features(sp)
    row["رنگ بند"] = norm_color(sp.get("رنگ بند", ""))
    row["رنگ قاب"] = norm_color(sp.get("رنگ قاب", ""))
    row["نوع موتور"] = _mapv("نوع موتور", sp)
    row["نوع شیشه"] = _mapv("نوع شیشه", sp)
    row["طرح بند"] = _mapv("طرح بند", sp)
    row["نوع قفل"] = _mapv("نوع قفل", sp)
    row["تقویم و نوع آن"] = _mapv("تقویم و نوع آن", sp)
    row["نگین"] = ""            # ثابتِ برند — اپراتور پر می‌کند
    row["گارانتی"] = ""          # ثابتِ برند
    row["گارانتی کننده در ایران"] = ""  # ثابتِ برند
    row["موارد گارانتی"] = "کارکرد موتور | مقاومت در برابر آب"  # ثابت برای همهٔ برندها (اپراتور)
    row["اصالت برند"] = _mapv("اصالت برند", sp) or _norm(sp.get("اصالت کشور برند", ""))
    row["کشور سازنده"] = _mapv("کشور سازنده", sp) or _norm(sp.get("اصالت کشور برند", ""))
    row["سایز قاب"] = _size(sp, "عرض قاب")
    row["ارتفاع قاب"] = _size(sp, "ارتفاع قاب")
    row["عرض بند"] = _size(sp, "عرض بند")
    row["وزن ساعت"] = _norm(sp.get("وزن ساعت", ""))
    return row
